import io
from datetime import datetime

import openpyxl
import streamlit as st

from modules.receita_financeira import (
    processar_receita_financeira,
    ColunaNaoEncontradaError,
    COLUNA_MONTANTE,
)


def _resumo_dinamica(bytes_xlsx: bytes):
    """Lê a aba 'Dinâmica' do arquivo já processado (sem fórmulas calculadas em
    memória, então soma direto a partir da aba 'Análise', que já tem PIS/COFINS
    como fórmula) para exibir métricas de conferência na tela."""
    wb = openpyxl.load_workbook(io.BytesIO(bytes_xlsx), data_only=False)
    ws = wb["Análise"]

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_montante = headers.get(COLUNA_MONTANTE)
    col_pis = headers.get("PIS")
    col_cofins = headers.get("COFINS")

    total_montante = 0.0
    total_pis = 0.0
    total_cofins = 0.0
    linhas = 0

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_montante).value
        if isinstance(v, (int, float)):
            total_montante += v
            total_pis += v * 0.0065
            total_cofins += v * 0.04
            linhas += 1

    return {
        "linhas": linhas,
        "total_montante": total_montante,
        "total_pis": total_pis,
        "total_cofins": total_cofins,
        "abas": wb.sheetnames,
    }


def _formatar_moeda(valor: float) -> str:
    texto = f"{valor:,.2f}"
    texto = texto.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {texto}"


def render(render_botao_voltar):
    render_botao_voltar()

    st.markdown(
        """
        <div class="vix-page-header">
            <div class="vix-page-icon">💰</div>
            <h2 class="vix-page-title">Receita Financeira</h2>
        </div>
        <p class="vix-page-sub">Envie o relatório de Receita Financeira e gere a planilha com PIS, COFINS e a tabela-resumo.</p>
        """,
        unsafe_allow_html=True,
    )

    passo_1_ok = "rf_resultado" in st.session_state
    classe_1 = "vix-step-done" if passo_1_ok else ""

    st.markdown(
        f"""
        <div class="vix-step {classe_1}">
            <div class="vix-step-num">{'✓' if passo_1_ok else '1'}</div>
            <div class="vix-step-text">Envie o Excel do relatório de Receita Financeira (apenas a aba com os dados brutos).</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    arquivo = st.file_uploader(
        "Excel da Receita Financeira",
        type=["xlsx", "xlsm"],
        label_visibility="collapsed",
        key="rf_uploader",
    )

    if arquivo is not None:
        novo_arquivo = st.session_state.get("rf_nome_arquivo") != arquivo.name
        if novo_arquivo or "rf_resultado" not in st.session_state:
            with st.spinner("Processando planilha: criando BD, Análise (PIS/COFINS) e Dinâmica..."):
                try:
                    resultado_bytes = processar_receita_financeira(arquivo)
                    st.session_state.rf_resultado = resultado_bytes
                    st.session_state.rf_nome_arquivo = arquivo.name
                    st.session_state.rf_erro = None
                except ColunaNaoEncontradaError as e:
                    st.session_state.rf_resultado = None
                    st.session_state.rf_erro = str(e)
                except Exception as e:
                    st.session_state.rf_resultado = None
                    st.session_state.rf_erro = (
                        f"Não foi possível processar o arquivo: {e}"
                    )

    erro = st.session_state.get("rf_erro")
    if erro:
        st.markdown(
            f"""
            <div class="vix-callout" style="background:#FBE9E9; color:#8A1F1F; border-color:#F3C9C9;">
                <strong>Não foi possível processar o arquivo.</strong><br>{erro}
            </div>
            """,
            unsafe_allow_html=True,
        )

    resultado = st.session_state.get("rf_resultado")

    passo_2_ok = resultado is not None
    classe_2 = "vix-step-done" if passo_2_ok else ""
    st.markdown(
        f"""
        <div class="vix-step {classe_2}">
            <div class="vix-step-num">{'✓' if passo_2_ok else '2'}</div>
            <div class="vix-step-text">PIS (0,65%) e COFINS (4%) calculados sobre o Montante e tabela-resumo gerada.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if resultado is not None:
        resumo = _resumo_dinamica(resultado)

        st.write("")
        c1, c2, c3 = st.columns(3)
        c1.metric("Linhas processadas", f"{resumo['linhas']:,}".replace(",", "."))
        c2.metric("Soma do Montante", _formatar_moeda(resumo["total_montante"]))
        c3.metric("PIS + COFINS", _formatar_moeda(resumo["total_pis"] + resumo["total_cofins"]))

        st.markdown(
            f"""
            <div class="vix-callout vix-callout-success">
                Planilha gerada com as abas <strong>{' · '.join(resumo['abas'][:3])}</strong>.
                A aba <strong>BD</strong> preserva os dados originais sem alterações.
            </div>
            """,
            unsafe_allow_html=True,
        )

        nome_saida = st.session_state.get("rf_nome_arquivo", "receita_financeira.xlsx")
        nome_base = nome_saida.rsplit(".", 1)[0] if "." in nome_saida else nome_saida
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        nome_download = f"{nome_base}_PIS_COFINS_{timestamp}.xlsx"

        st.download_button(
            label="⬇ Baixar planilha com PIS, COFINS e Dinâmica",
            data=resultado,
            file_name=nome_download,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        with st.expander("O que foi feito na planilha"):
            st.markdown(
                """
                - **BD** — cópia exata da aba original enviada, sem nenhuma alteração.
                - **Análise** — cópia de BD com duas colunas novas ao lado de
                  *Montante em moeda interna*: **PIS** (Montante × 0,65%) e
                  **COFINS** (Montante × 4%), calculadas por fórmula do Excel.
                - **Dinâmica** — tabela-resumo agrupada por *Empresa, Loc.negócios,
                  Conta, Centro de lucro, Segmento* e *Texto*, somando
                  *Montante em moeda interna*, *PIS* e *COFINS*.
                """
            )
