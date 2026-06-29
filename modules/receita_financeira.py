"""
Processamento do relatório de Receita Financeira (PIS/COFINS).

Regras de negócio (definidas em Automatizar.md):
  1. Aba 1 do arquivo original -> renomeada/copiada para "BD" (intocada, fonte segura).
  2. Cópia de BD -> aba "Análise", com duas novas colunas inseridas imediatamente
     à direita de "Montante em moeda interna":
        - PIS    = Montante * 0,65%
        - COFINS = Montante * 4%
  3. Nova aba "Dinâmica" contendo uma tabela dinâmica nativa do Excel com:
        Linhas:  Empresa, Loc.negócios, Conta, Centro de lucro, Segmento, Texto
        Valores: Soma de Montante em moeda interna, Soma de PIS, Soma de COFINS
"""

from __future__ import annotations

import io
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.pivot.cache import CacheDefinition, CacheField, SharedItems, CacheSource, WorksheetSource
from openpyxl.pivot.fields import Number, Text
from openpyxl.pivot.record import Record, RecordList
from openpyxl.pivot.table import (
    TableDefinition,
    Location,
    PivotField,
    RowColField,
    RowColItem,
    DataField,
    Index,
)

COLUNA_MONTANTE = "Montante em moeda interna"
ALIQUOTA_PIS = 0.0065
ALIQUOTA_COFINS = 0.04

CAMPOS_LINHA_DINAMICA = [
    "Empresa",
    "Loc.negócios",
    "Conta",
    "Centro de lucro",
    "Segmento",
    "Texto",
]
CAMPOS_VALOR_DINAMICA = [COLUNA_MONTANTE, "PIS", "COFINS"]


class ColunaNaoEncontradaError(Exception):
    pass


def _encontrar_coluna(ws, nome_coluna: str) -> int:
    """Retorna o índice (1-based) da coluna cujo cabeçalho (linha 1) bate com nome_coluna."""
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor is not None and str(valor).strip() == nome_coluna:
            return col
    raise ColunaNaoEncontradaError(
        f"Coluna '{nome_coluna}' não encontrada na primeira linha da planilha."
    )


def _copiar_estilo(origem_cell, destino_cell):
    if origem_cell.has_style:
        destino_cell.font = copy(origem_cell.font)
        destino_cell.border = copy(origem_cell.border)
        destino_cell.fill = copy(origem_cell.fill)
        destino_cell.number_format = copy(origem_cell.number_format)
        destino_cell.protection = copy(origem_cell.protection)
        destino_cell.alignment = copy(origem_cell.alignment)


def _duplicar_aba(wb, ws_origem, novo_nome: str):
    """Duplica completamente uma aba (valores, estilos, larguras de coluna) com novo nome."""
    ws_nova = wb.copy_worksheet(ws_origem)
    ws_nova.title = novo_nome
    return ws_nova


def _construir_aba_analise(wb, ws_bd):
    """Cria a aba 'Análise' a partir de 'BD', inserindo as colunas PIS e COFINS
    imediatamente à direita de 'Montante em moeda interna'."""
    ws_analise = _duplicar_aba(wb, ws_bd, "Análise")

    col_montante = _encontrar_coluna(ws_analise, COLUNA_MONTANTE)
    col_pis = col_montante + 1
    col_cofins = col_montante + 2

    ws_analise.insert_cols(col_pis, amount=2)

    letra_montante = get_column_letter(col_montante)

    ws_analise.cell(row=1, column=col_pis, value="PIS")
    ws_analise.cell(row=1, column=col_cofins, value="COFINS")

    header_ref_col = col_montante if col_montante != col_pis else col_pis
    cabecalho_modelo = ws_analise.cell(row=1, column=header_ref_col)
    for c in (col_pis, col_cofins):
        _copiar_estilo(cabecalho_modelo, ws_analise.cell(row=1, column=c))

    ultima_linha = ws_analise.max_row
    for row in range(2, ultima_linha + 1):
        cel_montante = ws_analise.cell(row=row, column=col_montante)
        if cel_montante.value is None:
            continue
        ref = f"{letra_montante}{row}"
        cel_pis = ws_analise.cell(row=row, column=col_pis, value=f"={ref}*0.65%")
        cel_cofins = ws_analise.cell(row=row, column=col_cofins, value=f"={ref}*4%")
        cel_pis.number_format = "#,##0.00"
        cel_cofins.number_format = "#,##0.00"

    largura_origem = ws_bd.column_dimensions[letra_montante].width
    for c in (col_pis, col_cofins):
        letra = get_column_letter(c)
        ws_analise.column_dimensions[letra].width = largura_origem or 14

    return ws_analise, col_montante, col_pis, col_cofins


def _construir_dinamica_nativa(wb, ws_analise, col_montante, col_pis, col_cofins):
    """Cria a aba 'Dinâmica' com uma tabela dinâmica NATIVA do Excel (objeto
    PivotTable real, com painel de campos, arrastar-e-soltar, etc.), agrupada
    por Empresa, Loc.negócios, Conta, Centro de lucro, Segmento e Texto, somando
    Montante em moeda interna, PIS e COFINS — replicando a estrutura usada no
    arquivo de referência da VIXPAR (BD -> Análise -> Dinâmica)."""
    ultima_linha = ws_analise.max_row
    ultima_coluna = ws_analise.max_column

    idx_por_nome = {}
    for col in range(1, ultima_coluna + 1):
        nome = ws_analise.cell(row=1, column=col).value
        if nome:
            idx_por_nome[str(nome).strip()] = col

    colunas_linha = [c for c in CAMPOS_LINHA_DINAMICA if c in idx_por_nome]
    colunas_valor = [c for c in CAMPOS_VALOR_DINAMICA if c in idx_por_nome]

    if not colunas_linha or not colunas_valor:
        # Sem campos suficientes para montar a tabela dinâmica: não falha o
        # processamento todo, apenas não cria a aba Dinâmica.
        return None

    # Conjunto de colunas que entram no cache da pivot: primeiro os campos de
    # linha, depois os campos de valor (ordem que vamos usar nos índices fld).
    colunas_cache = colunas_linha + colunas_valor
    indices_origem = [idx_por_nome[c] for c in colunas_cache]

    cache_fields = [CacheField(name=c, sharedItems=SharedItems()) for c in colunas_cache]

    montante_idx = idx_por_nome[COLUNA_MONTANTE]

    records = []
    for r in range(2, ultima_linha + 1):
        montante = ws_analise.cell(row=r, column=montante_idx).value
        if not isinstance(montante, (int, float)):
            continue
        valores_linha = []
        for c in colunas_linha:
            v = ws_analise.cell(row=r, column=idx_por_nome[c]).value
            valores_linha.append("" if v is None else str(v))
        # Calculamos PIS/COFINS em Python (o cache da pivot exige valores já
        # resolvidos; a aba Análise mantém as fórmulas para auditoria visual).
        valores_calculo = {
            COLUNA_MONTANTE: montante,
            "PIS": montante * ALIQUOTA_PIS,
            "COFINS": montante * ALIQUOTA_COFINS,
        }

        fields = []
        for c in colunas_linha:
            v = ws_analise.cell(row=r, column=idx_por_nome[c]).value
            fields.append(Text(v="" if v is None else str(v)))
        for c in colunas_valor:
            fields.append(Number(v=valores_calculo[c]))

        records.append(Record(_fields=fields))

    if not records:
        return None

    record_list = RecordList(r=records)

    n_linha = len(colunas_linha)
    n_valor = len(colunas_valor)
    n_total = n_linha + n_valor

    cache_source = CacheSource(
        type="worksheet",
        worksheetSource=WorksheetSource(
            ref=f"A1:{get_column_letter(ultima_coluna)}{ultima_linha}",
            sheet=ws_analise.title,
        ),
    )
    cache = CacheDefinition(
        cacheSource=cache_source,
        cacheFields=cache_fields,
        recordCount=len(records),
        refreshOnLoad=True,
    )
    cache.records = record_list

    pivot_fields = []
    for i in range(n_total):
        if i < n_linha:
            pivot_fields.append(PivotField(axis="axisRow", showAll=False, defaultSubtotal=False))
        else:
            pivot_fields.append(PivotField(dataField=True, showAll=False))

    row_fields = [RowColField(x=i) for i in range(n_linha)]
    row_items = [RowColItem(x=[Index(v=0)])]

    col_fields = [RowColField(x=-2)]
    col_items = [RowColItem(x=[Index(v=0)])]
    for i in range(1, n_valor):
        col_items.append(RowColItem(i=i, x=[Index(v=i)]))

    data_fields = [
        DataField(name=f"Soma de {c}", fld=n_linha + k, baseField=0, baseItem=0)
        for k, c in enumerate(colunas_valor)
    ]

    ultima_col_pivot = get_column_letter(n_total)
    pivot_table = TableDefinition(
        name="TabelaDinamicaPISCOFINS",
        cacheId=1,
        dataCaption="Valores",
        pivotFields=pivot_fields,
        rowFields=row_fields,
        rowItems=row_items,
        colFields=col_fields,
        colItems=col_items,
        dataFields=data_fields,
        location=Location(
            ref=f"A3:{ultima_col_pivot}10",
            firstHeaderRow=1,
            firstDataRow=2,
            firstDataCol=n_linha,
        ),
    )
    pivot_table.cache = cache

    ws_dinamica = wb.create_sheet("Dinâmica")
    ws_dinamica.add_pivot(pivot_table)
    return ws_dinamica


def processar_receita_financeira(arquivo_excel) -> bytes:
    """
    Recebe um arquivo Excel (caminho, bytes ou file-like) com o relatório de
    Receita Financeira e devolve, em memória, os bytes do arquivo já processado:
      - BD: cópia intocada da primeira aba do arquivo de origem
      - Análise: cópia de BD + colunas PIS e COFINS
      - Dinâmica: tabela dinâmica (Empresa, Loc.negócios, Conta, Centro de lucro,
        Segmento, Texto) somando Montante, PIS e COFINS

    Caso o arquivo enviado já possua abas chamadas "BD", "Análise" ou "Dinâmica"
    (por exemplo, um arquivo já processado anteriormente), elas são removidas
    antes do reprocessamento para evitar duplicidade (Análise1, Dinâmica1, etc.).
    """
    wb = openpyxl.load_workbook(arquivo_excel, data_only=False)

    ws_original = wb.worksheets[0]

    for nome_reservado in ("BD", "Análise", "Dinâmica"):
        if nome_reservado in wb.sheetnames and wb[nome_reservado] is not ws_original:
            del wb[nome_reservado]

    ws_original.title = "BD"
    ws_bd = wb["BD"]

    ws_analise, col_montante, col_pis, col_cofins = _construir_aba_analise(wb, ws_bd)

    _construir_dinamica_nativa(wb, ws_analise, col_montante, col_pis, col_cofins)

    ordem_desejada = ["BD", "Análise", "Dinâmica"]
    outras_abas = [n for n in wb.sheetnames if n not in ordem_desejada]
    wb._sheets = [wb[n] for n in ordem_desejada] + [wb[n] for n in outras_abas]

    wb.active = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
